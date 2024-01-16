import os
import re
import time
from openpyxl import *
import dbms_function

db_path = 'data/'
# view_path = 'view/'
user = ''
using_dbname = ''
using_db = Workbook()





def help():
    """
    打印帮助信息
    :return:
    """
    print("""
    右下角输入指令，submit等于回车

    点击数据库按钮刷新指令

    点击全选选择所有数据库

    点击加载加载数据

    加载完成后可在data中查看每个数据库以及sheet

    ## 登录管理员
    username:admin
    username:admin

    ## 创建数据库
    create database {database_name}
    eg.: create database test_db

    ## 删除数据库
    drop database {database_name}
    eg.: drop database test_db

    ## 使用数据库
    use database {database_name}
    eg.: use database test_db

    ## 创建表
    create table {table_name} ({column_name} {data_type} {PK,null...},{column_name} {data_type} {PK,null...}...)
    eg.: create table test (v1 int PK null,v2 int)

    ## 删除表
    drop table {table_name}
    eg.: drop table test

    ## 添加字段
    alter {table_name} add ({column_name} {data_type} {PK,null...})
    eg.: alter test add (v3 int)

    ## 删除字段
    alter {table_name} drop ({column_name})
    eg.: alter test drop (v3)

    ## 修改字段
    alter {table_name} modify {alter_field_name} ({column_name} {data_type} {PK,null...}) 
    eg.: alter test modify v1 (v3 int PK null)
    
    ## 记录插入
    insert into {table_name} {column_name=value,column_name=value,...)
    eg.: insert into test v1=1,v2=2

    ## 记录插入（多重）
    insert into {table_name} {column_name=value,column_name=value,...&column_name=value,column_name=value,...)
    eg.: insert into test v3=2,v2=4&v3=3,v2=5

    ## 记录删除
    delete on {table_name} where {column_name=value或column_name>value或column_name<value}
    eg.: delete on test where v3=1

    ## 记录删除（多重）
    delete on {table_name} where {column_name=value或column_name>value或column_name<value&column_name=value或column_name>value或column_name<value&..}
    eg.: delete on test where v3=1&v2=2

    ## 记录修改
    update {table_name} set column_name=value,column_name=value,... where {column_name=value或column_name>value或column_name<value（可多重）}
    eg.: update test set v3=4,v2=3 where v3=2

    ## 选择全部
    select * from {table_name}
    eg.: select * from test

    ## 选择指定列
    select {column_name} from {table_name}
    eg.:select v3 from test

    ## 选择where条件
    select * 或{column_name} from {table_name} where {column_name=value或column_name>value或column_name<value（可多重）}
    eg.: select * from test where v3=4

    ## 注册用户
    signup {username} {password}
    eg.: signup admin admin

    ## 读取脚本
    load {script_name}
    eg.: load test.txt

    ## 创建视图
    create view {view_name} as select * 或{column_name} from {table_name}
    eg.: create view test as select * from test

    ## 赋予权限
    grant {action} on {database_name} for {username}
    eg.: grant select on test_db for aaa

    ## 收回权限
    revoke {action} on {database_name} for {username}
    eg.: revoke select on test_db for aaa

    """)


# 使用数据库
def use_db(dbname):
    global using_dbname
    global using_db
    # 数据库不存在
    if os.path.exists(db_path + dbname + '.xlsx'):
        if dbms_function.check_permission(user, dbname, 'use'):
            using_dbname = dbname
            print(dbname + "数据库已使用.")
            using_db = load_workbook(db_path + dbname + '.xlsx')
        else:
            print("你没有权限使用该数据库,请使用admin账户赋予权限.")
    else:
        print("数据库不存在")

    # 显示所有数据库


def show_db():
    print("All database:")
    dbs = os.listdir(db_path)  # 第二种方法，从保存数据库信息的库中查询
    for db in dbs:
        if '.DS' not in db and db != 'index':
            print("[*] " + db[:-5])


# 创建数据库
def creat_db(dbname):
    dbpath = 'data/' + dbname + '.xlsx'
    database = Workbook()
    database.save(dbpath)
    dbms_function.create_tb_in_tbinfo(dbname)
    print(u"数据库创建操作执行成功")


def get_command():
    """
    从控制台获取命令
    :return: None
    """
    command = input("[👉]> ") if not using_dbname else input("[{}🚩]> ".format(using_dbname))
    # command = command.lower()
    # print command
    return command.strip()


def Initialization():
    if not os.path.exists(db_path):
        os.mkdir(db_path)
    if not os.path.exists("data/table_information.xlsx"):
        Workbook().save("data/table_information.xlsx")
    if os.path.exists("data/system.xlsx"):
        print("Initializating......")
    else:
        dbms_function.creat_db('system')
    db = load_workbook("data/system.xlsx")
    permission_tb_col = ['database char[50] pk unique', 'select char', 'insert char', 'delete char', 'update char']
    dbms_function.creat_table('permission', db, 'system', permission_tb_col)


def query(sql, tag=''):
    sql_word = sql.split(" ")
    global using_dbname
    global using_db
    if len(sql_word) < 2:
        print("[!] Wrong query!")
        return
    operate = sql_word[0].lower()
    # 使用数据库
    if operate == 'use':
        if sql_word[1] == 'database':
            try:
                use_db(sql_word[2])
            except:
                print("[!]Error")
        else:
            print("[!]Syntax Error.\neg:>use database dbname")
    # 创建数据库、表、视图、索引
    elif operate == 'create':
        if sql_word[1] == 'database':
            try:
                creat_db(sql_word[2])
            except:
                print("[!]Create Error")
        elif sql_word[1] == 'table':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            print(columns_list, using_dbname)
            try:
                dbms_function.creat_table(sql_word[2], using_db, using_dbname, columns_list)
                using_db = load_workbook(db_path + using_dbname + '.xlsx')
            except:
                print("[!]Error")
        #创建视图
        elif sql_word[1] == 'view':
            #若sql_word[2]存在
            if sql_word[2]:
                viewname = sql_word[2]
                if sql_word[3] == 'as' and sql_word[4] == 'select':
                    sql = sql_word[5:]
                    dbms_function.create_view(viewname, sql,using_db)
                else:
                    print("[!]Syntax Error.\neg:>create view viewname as select * from table")
            else:
                print("[!]Syntax Error.\neg:>create view viewname as select * from table")

        elif sql_word[1] == 'index':
            return
        else:
            print("[!]Syntax Error.")
    # 删除数据库或表
    elif operate == 'drop':
        if sql_word[1] == 'database':
            try:
                dbms_function.drop_db(sql_word[2])
            except:
                print("[!]Error")
        if sql_word[1] == 'table':
            try:
                dbms_function.drop_table(sql_word[2], using_dbname, using_db)
                #若数据库还存在
                if os.path.exists(db_path + using_dbname + '.xlsx'):
                    using_db = load_workbook(db_path + using_dbname + '.xlsx')
            except:
                print("[!]Error")
    # 字段操作alter
    elif operate == 'alter':
        # 添加字段
        if sql_word[2] == 'add':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            try:
                dbms_function.add_field(tbname=sql_word[1], columns_list=columns_list, using_dbname=using_dbname,using_db=using_db)
                using_db = load_workbook(db_path + using_dbname + '.xlsx')
            except:
                print("[!]Error")
        # 删除字段
        elif sql_word[2] == 'drop':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            try:
                dbms_function.drop_field(tbname=sql_word[1], columns_list=columns_list, using_dbname=using_dbname,using_db=using_db)
                using_db = load_workbook(db_path + using_dbname + '.xlsx')
            except:
                print("[!]Error")
        # 修改字段
        elif sql_word[2] == 'modify':
            columns_list = re.findall('\((.*)\)', sql)[0].split(',')
            try:
                dbms_function.modify_field(tbname=sql_word[1], alterFieldName=sql_word[3], columns_list=columns_list,using_dbname=using_dbname, using_db=using_db)
                using_db = load_workbook(db_path + using_dbname + '.xlsx')
            except:
                print("[!]Error")
    # 选择操作select
    elif operate == 'select':
        pos = 0
        for i in range(len(sql_word)):
            if '(' in sql_word[i] and 'select' in sql_word[i]:
                pos = i
        if pos == 3:
            sql2 = sql_word[3][1:-1]
            query(sql2, tag='nesting')
            sql_word[3] = 'tmp'
            sql = ' '.join(sql_word)

        columns = sql_word[1]
        table_name = sql_word[3]
        if len(sql_word) > 4:
            # try:
            limit = sql_word[5].split()
            predicate = 'and'
            symbol = '='
            if ',' in sql_word[5]:
                limit = sql_word[5].split(',')
                predicate = 'and'
            elif '|' in sql_word[5]:
                limit = sql_word[5].split('|')
                predicate = 'or'
            elif '>' in sql_word[5]:
                # limit = sql_word[5].split()
                symbol = '>'
            elif '<' in sql_word[5]:
                # limit = sql_word[5].split()
                symbol = '<'
            elif len(sql_word) > 6:
                if sql_word[6] == 'in':
                    limit = [sql_word[5] + '=' + sql_word[7]]
                    predicate = 'in'
                if sql_word[6] == 'like':
                    limit = [sql_word[5] + '=' + sql_word[7]]
                    predicate = 'like'
            # except:
            # limit = [].append(sql_word[5])
            # print limit
            for i in range(len(limit)):
                limit[i] = limit[i].split(symbol)
            limit = dict(limit)
            return dbms_function.select(columns, table_name, using_dbname, using_db, limit, predicate=predicate,
                                        symbol=symbol, tag=tag)
        else:  # 没where的情况
            return dbms_function.select(columns, table_name, using_dbname, using_db, tag=tag)
    # 授予权限
    elif operate == 'grant':
        if user != 'admin':
            return False
        dbms_function.set_permission(sql_word[5], sql_word[3], sql_word[1])
    # 取消权限
    elif operate == 'revoke':
        if user != 'admin':
            return False
        dbms_function.del_permission(sql_word[5], sql_word[3], sql_word[1])
    # 插入数据
    elif operate == 'insert':  # INSERT INTO table_name col1=val1,col2=val2&col3=val3,col4=val4
        table_name = sql_word[2]
        """
        #INSERT INTO table_name (select x from xx)
        sql2 = re.findall('\((.*)\)')[0]
        query(sql2,tag='insert')
        """
        multiFlag = False

        columns_list = []
        if '&' in sql:
            multiFlag = True
            cols = sql_word[3].split('&')  # [{xx},{xx}] 多组
            for i in range(len(cols)):
                cols[i] = cols[i].split(',')
            for i in range(len(cols)):
                for j in range(len(cols[i])):
                    cols[i][j] = cols[i][j].split('=')
        else:
            cols = sql_word[3].split(',')
            for i in range(len(cols)):
                cols[i] = cols[i].split('=')
        dbms_function.insert_record(table_name, using_db, using_dbname, cols, multiFlag)
    # 删除记录
    elif operate == 'delete':
        table_name = sql_word[2]
        if 'where' == sql_word[3]:
            condition_list = sql_word[4].split(',')
            dbms_function.delete_record(table_name, using_db, using_dbname, condition_list)
        else:
            print("[!]Syntax Error.")

    # 修改记录
    elif operate == 'update':
        table_name = sql_word[1]
        # 处理set后的=赋值部分
        if 'set' == sql_word[2]:
            multiFlag = False
            columns_list = []
            cols = sql_word[3].split(',')
            for i in range(len(cols)):
                cols[i] = cols[i].split('=')
        else:
            print("[!]Syntax Error.")
        # 处理where后的条件部分
        if 'where' == sql_word[4]:
            condition_list = sql_word[5].split(',')
        else:
            print("[!]Syntax Error.")
        # 调用函数update
        dbms_function.update_record(table_name, using_db, using_dbname, cols, condition_list, multiFlag)

    #注册用户
    elif operate == 'signup':
        if user != 'admin':
            print("[!]You are not admin.")
            return False
        #若sql_word[1]不存在
        try:
            dbms_function.signup(sql_word[1], sql_word[2])
        except:
            print("[!]Syntax Error.")

    # 帮助指令
    elif operate == 'help':
        if sql_word[1] == 'database':
            dbms_function.show_db()
        if sql_word[1] == 'table':
            usdbnm = using_dbname
            use_db('table_information')
            # 若sql_word[2]存在，则指定表
            if len(sql_word) > 2 and sql_word[2] != '':
                tbname = sql_word[2]
                dbms_function.select('*', usdbnm, {'table': tbname})
            else:
                print('[!]Syntax Error.\neg:>help table table_name')
        if sql_word[1] == 'view':
            view_name = sql_word[2]
            use_db('view')
            dbms_function.select('sql', 'sql', {'viewnamw': view_name})
        if sql_word[1] == 'index':
            print("All Index:")
            indexs = os.listdir('data/index/')  # 第二种方法，从保存数据库信息的库中查询
            for index in indexs:
                if '.DS' not in index:
                    print("[*] " + index[:-5])
    
    # 读取脚本
    elif operate == 'load':
        #弹出窗口按行读取txt
        file_name = sql_word[1]
        file_path = 'data/script/' + file_name + '.txt'
        if os.path.exists(file_path):
            #按行读取txt文件
            f = os.open(file_path, os.O_RDONLY)
            lines = os.read(f, os.path.getsize(file_path)).decode('utf-8').split('\n')
            for line in lines:
                query(line)
    else:
        print("[!]Syntax Error.")


def run():
    # Initialization()
    global user
    # welcome()
    user = dbms_function.login(user)
    while True:
        command = get_command()
        # print command
        if command == 'quit' or command == 'exit':
            print("[🍻] Thanks for using Mini DBMS. Bye~~")
            exit(0)
        elif command == 'help':
            help()
        else:
            query(command)


# 若没有system和table_information库，则使用此方法创建创建
# if __name__ == '__main__':
# Initialization()
# run()
# 登录
def userLogin(username, password, flagFirst, flagLogin):
    global user
    user, flagFirst, flagLogin = dbms_function.login(user, username, password, flagFirst, flagLogin)
    return flagFirst, flagLogin


def interpreter(command):
    if command == 'quit' or command == 'exit':
        print("[🍻] Thanks for using Mini DBMS. Bye~~")
        exit(0)
    elif command == 'help':
        help()
    else:
        query(command)
    #将command写入log.txt
    f = os.open('data/log.txt', os.O_RDWR | os.O_APPEND)
    #写入时间+command
    byte_str = ((time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())) + ' ' + command + '\n').encode('utf-8')
    os.write(f, byte_str)
    os.close(f)


